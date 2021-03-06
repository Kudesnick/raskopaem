# -*- coding: utf-8 -*-

import sys
import os
import random
from timeit import default_timer
from pathlib import Path
from openpyxl import load_workbook
from openpyxl import Workbook
from sys import argv
import pkg_resources
from subprocess import call

time_start = default_timer()

print('''
Typology parser 2019.
author: Stulov Tikhon (aka Kudesnick)
mailto: kudesnick@inbox.ru
git: https://github.com/Kudesnick/raskopaem.git
''')

# constants
curr_encoding = 'windows-1251'
path_input = 'input'
table_ext = '.xlsx'
img_ext = '.png'
img_size = (32, 32) # inches
typo_f_name = 'table_tipology'
lists_f_name = 'lists'
out_f_name = 'lists_out'
log_f_name = 'log.txt'
typo_splitter = '|'

def err(_str):
    print(''.join(['Error. ', _str]))
    sys.exit()

def flag_is_set(_flag : str):
    for i in range(1, len (argv)):
        if argv[i] == _flag:
            return True
    return False

# upgrade packages
if flag_is_set('-u'):
    print('Packages updating..')
    packages = [dist.project_name for dist in pkg_resources.working_set]
    call("pip install --upgrade " + ' '.join(packages), shell=True)

# drawing object
fig = None

if flag_is_set('-g'):
    import matplotlib.pyplot as plt
    fig = plt.figure(str(Path(out_f_name).with_suffix(img_ext)), figsize = img_size)

# convert xmlx to dict
def exel_to_dict(_path : Path):
    
    result = dict()

    if not _path.is_file():
        err(''.join(['File "', str(_path), '" not exists.']))

    wb = load_workbook(_path)

    for sht in wb.sheetnames:
        print('{}..'.format(str(sht)))
        # работаем отдельно с каждым листом
        result.update({sht: []})
        ws = wb[sht]

        # инициализируем поля
        keys = []
        for c in range(1, ws.max_column + 1):
            if ws.cell(1, c).value != '':
                keys.append(ws.cell(1, c).value)
            else:
                break
        # собираем данные из строк
        for r in range(2, ws.max_row + 1):
            values = [ws.cell(row=r,column=i).value for i in range(1,len(keys)+1)]
            if ''.join(str(i) for i in values) != '':
                result[sht].append(dict(zip(keys, values)))
            else:
                break
    return result

print('typology parsing..')

# open temlate of typology
typo_path = Path(path_input, typo_f_name).with_suffix(table_ext)
typo_obj = exel_to_dict(typo_path)

# convert typology objects
for i in typo_obj['table_tipology']:
    i['variants'] = list(map(lambda x: str(x).strip().lower(), i['variants'].split(typo_splitter)))

# convert settings sheet
sett = {row['param']: row['value'] for row in typo_obj['settings']}

print('lists parsing..')

# open main lists
lists_path = Path(path_input, lists_f_name).with_suffix(table_ext)
lists_obj = exel_to_dict(lists_path)

# return typology from string
def get_typo(_str : str, strip = ' '):
    _str = _str.strip().lower()
    _words = []
    if sett['word_only'] == 'y':
        _words = list(map(lambda x: str(x).strip(strip), _str.split()))
    for i in typo_obj['table_tipology']:
        for j in i['variants']:
            result = True
            for k in j.split():
                if len(k) > 2:
                    if sett['word_only'] == 'y':
                        try:
                            _words.index(k)
                        except:
                            result = False
                    elif _str.find(k) < 0:
                        result = False
                if result == False:
                    break
            if result == True:
                return i['alias']
    return None

# get horizon
def get_horizon(_str : str):
    _str = _str.strip().lower()
    for i in typo_obj['horizon']:
        if _str == str(i['horizon']).strip().lower():
            return [int(i['min']), int(i['max'])]
    return get_horizon('default')

# random generator init
random.seed(1024)

# open logfile
logfile = open(Path(path_input, log_f_name), 'w', encoding = curr_encoding)

# add typology and coordinates
print('typologies and coordinates adding..')

min_x = -1
min_y = -1
max_x = -1
max_y = -1

for year, rows in lists_obj.items():
    print('{}..'.format(str(year)))

    prev_ltr = None
    prev_num = None
    prev_hor = None
    prev_typo = None
    prev_locate = None
    prev_year = None
    q_ltrs = None

    description_str = None
    horizon_str = None
    quad_letter_str = None
    quad_num_str = None

    for n, i in enumerate(rows):
        if sett['split_coord'] == 'y':
            i.update({'X': None, 'Y': None, 'Z': None})

        first_row = bool(n == 0)
        err_str = 'Lists error! page {p}, row {n} '.format(p = str(year), n = str(n + 2))
        
        # set typology ->
        if i['description'] != None:
            description_str = i['description']
            prev_typo = get_typo(str(i['description']), sett['split_symbols'])
        
        if prev_typo == None:
            if i['description'] != None or first_row:
                print('{}description is invalid! "{}"'.format(err_str, str(i['description'])), file = logfile)
            i['type'] = 'Error!'
        else:
            i['type'] = prev_typo
            i['description'] = description_str
        # set typology <-
        
        #set coord ->
        # add locale
        if i['locate'] != None:
            try:
                prev_locate = int(i['locate'])
            except:
                prev_locate = None
            if prev_locate != None:
                q_ltrs = None 
                for loc in typo_obj['locate']:
                    if int(loc['number']) == prev_locate:
                        q_ltrs = str(loc['letters'])
                        break
                if q_ltrs == None:
                    prev_locate = None
        
        if prev_locate == None:
            if i['locate'] != None or first_row:
                print('{}locate is invalid! "{}"'.format(err_str, str(i['locate'])), file = logfile)
        
        elif i['locate'] == None:
            i['locate'] = prev_locate

        # add quad letter
        if i['quad_letter'] != None:
            quad_letter_str = i['quad_letter']
            ql_st = str(i['quad_letter']).strip().lower()
            if len(ql_st) < 1:
                prev_ltr = None
            elif q_ltrs != None:
                if q_ltrs.find(ql_st) < 0: ql_st = ql_st[::-1] # slice string
                prev_ltr = [q_ltrs.find(ql_st)]
                prev_ltr.append(prev_ltr[0] + len(ql_st) - 1)
                if prev_ltr[0] < 0 or prev_ltr[1] < 0:
                    prev_ltr = None
            else:
                prev_ltr = None
        
        if prev_ltr == None:
            if i['quad_letter'] != None or first_row:
                print('{}quad letter is invalid! "{}"'.format(err_str, str(i['quad_letter'])), file = logfile)
        
        elif i['quad_letter'] == None: 
            i['quad_letter'] = quad_letter_str

        # add quad number
        if i['quad_num'] != None:
            quad_num_str = i['quad_num']
            qn_ls = str(i['quad_num']).split('-')
            if 1 <= len(qn_ls) <= 2:
                try:
                    prev_num = [int(str(qn_ls[0]).strip().lower())]
                    if len(qn_ls) < 2:
                        prev_num.append(prev_num[0])
                    else:
                        prev_num.append(int(str(qn_ls[1]).strip().lower()))
                    if (prev_num[0] > prev_num[1]):
                        prev_num[1], prev_num[0] = prev_num[0], prev_num[1]
                except:
                    prev_num = None
            else:
                prev_num = None
        
        if prev_num == None:
            if i['quad_num'] != None or first_row:
                print('{}quad number is invalid! "{}"'.format(err_str, str(i['quad_num'])), file = logfile)

        elif i['quad_num'] == None:
            i['quad_num'] = quad_num_str

        # add horizon
        if i['horizon'] != None:
            horizon_str = i['horizon']
            prev_hor = get_horizon(str(i['horizon']))
        
        if prev_hor == None:
            if i['horizon'] != None or first_row:
                print('{}horizon is invalid! "{}"'.format(err_str, str(i['horizon'])), file = logfile)
        elif i['horizon'] == None:
            i['horizon'] = horizon_str

        # set coord as is (relative quad a0), cm
        if prev_locate != None and prev_ltr != None and prev_num != None and prev_hor != None:
            
            offset = 0
            for loc in typo_obj['locate']:
                if int(loc['number']) == prev_locate:
                    break
                else:
                    offset = offset + len(str(loc['letters']))

            q = int(sett['quad_size'])
            mul_XY = 1 / float(sett['step_XY'])
            mul_Z  = 1 / float(sett['step_Z'])
            x = random.randint(prev_ltr[0] * q * mul_XY, (prev_ltr[1] * q + q) * mul_XY) + offset * q * mul_XY
            y = random.randint(prev_num[0] * q * mul_XY, (prev_num[1] * q + q) * mul_XY)
            z = random.randint(prev_hor[0] * mul_Z, prev_hor[1] * mul_Z)
            if sett['neg_Z'] == 'y':
                z = 0 - z

            if mul_XY != 1:
                x = float(x) / mul_XY
                y = float(y) / mul_XY

            if mul_Z != 1:
                z = float(z) / mul_Z

            if (min_x == -1 or x < min_x): min_x = x
            if (min_y == -1 or y < min_y): min_y = y
            if (max_x == -1 or x > max_x): max_x = x
            if (max_y == -1 or y > max_y): max_y = y

            i['coord'] = '{}:{}:{}'.format(x, y, z)
            
            if sett['split_coord'] == 'y':
                i['X'] = x
                i['Y'] = y
                i['Z'] = z

        else:
            i['coord'] = 'Error!'
        # set coord <-

        # flood void fields ->
        if i['year']   != None:
            prev_year = i['year']
        else:
            i['year'] = prev_year
        # flood void fields <-

        # coorect numbers ->
        if i['number'] != None:
            try:
                int(i['number'])
            except:
                i['number'] = str(i['number']).replace('\r', ' ').replace('\n', ' ')
        # coorect numbers <-

logfile.close()

# mirror coordinates
if (sett['mirr_x'] == 'y' or sett['mirr_y'] == 'y'):
    print('mirroring..')
    for year, rows in lists_obj.items():
        print('{}..'.format(str(year)))
        for n, i in enumerate(rows):
            if (i['X'] != None and sett['mirr_x'] == 'y'):
                i['X'] = max_x - int(i['X']) + min_x
            if (i['Y'] != None and sett['mirr_y'] == 'y'):
                i['Y'] = max_y - int(i['Y']) + min_y
            if (i['X'] != None and i['Y'] != None and i['Z'] != None):
                i['coord'] = '{}:{}:{}'.format(i['X'], i['Y'], i['Z'])

# create scatter
if fig != None:
    print('scatter create..')
    for year, rows in lists_obj.items():
        print('{}..'.format(str(year)))
        for n, i in enumerate(rows):
            if (i['X'] != None and i['Y'] != None):
                plt.scatter(i['X'], i['Y'])

# create new lists and save
print('saving results..')

wr_wb = Workbook()

for year, rows in lists_obj.items():
    print('{}..'.format(str(year)))
    sheet = wr_wb.create_sheet(title = str(year))
    sheet.append(list(rows[0].keys()))
    for r in rows:
        sheet.append(list(r.values()))

while True:
    fpath = Path(path_input, out_f_name).with_suffix(table_ext)
    try:
        wr_wb.save(fpath)
        break
    except:
        input('File "{}" access denied. May be this file is opened. Close file ant try again (press Enter).'.format(fpath))

# create img
if fig != None:
    print('image creating..')

    plt.savefig(Path(path_input, out_f_name).with_suffix(img_ext), fmt=img_ext.lstrip('.'))

print('complete!')
print('{} sec'.format(str(default_timer() - time_start)))

if fig != None:
    plt.show()
